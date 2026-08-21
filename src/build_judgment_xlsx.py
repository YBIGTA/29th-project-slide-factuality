#!/usr/bin/env python3
"""판정 결과(claim · 분류 · 근거 문장)를 덱마다 시트 하나로 엑셀에 모은다.

    python src/build_judgment_xlsx.py <판정폴더> -o 판정결과.xlsx

판정 파일은 덱마다 두 개다.

  {doc_id}.tsv         claim_id · 분류 · 근거참조
  {doc_id}.extra.tsv   (선택) 참조키 · 근거 문장 · 비고

`근거참조` 는 원문 문장 id(`s042`) 거나, extra.tsv 에 정의한 키다.
표는 정제 텍스트에서 걷혀 나가므로 PDF 에서 직접 옮긴 값을 extra 로 넣는다.

`근거 문장(맥락 포함)` 은 앞뒤 한 문장씩 붙인 것이다. 판정 근거가 되는
문장만 떼어 놓으면 지시어("이는", "이 경우")가 무엇을 가리키는지 사라져서
사람이 검수할 때 되짚기 어렵다.
"""
from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent))
from split_deck_claims import build as split_claims

COLUMNS = ["claim_id", "슬라이드", "claim", "분류", "근거 문장", "근거 문장 (맥락 포함)", "근거 위치"]
WIDTHS = [12, 8, 60, 10, 60, 78, 12]
CONTEXT = 1          # 앞뒤로 몇 문장까지 붙일지

LABEL_FILL = {"근거 있음": "BBEDC3", "무근거": "FDB9C3", "모순": "FFE88B", "Benign": "D0D0D0"}
HEADER_FILL = "355BB7"
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_sents(root: Path, doc_id: str) -> dict[str, tuple[str, int]]:
    """sent_id -> (본문, 순서). 순서가 있어야 앞뒤 문장을 붙일 수 있다."""
    path = root / "spans" / "sents" / f"{doc_id}.jsonl"
    out = {}
    for i, line in enumerate(path.open(encoding="utf-8")):
        if line.strip():
            r = json.loads(line)
            out[r["sent_id"]] = (r["text"], i)
    return out


def read_tsv(path: Path) -> list[list[str]]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8", newline="") as f:
        return [r for r in csv.reader(f, delimiter="\t") if r and r[0].strip()]


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("judgments", type=Path, help="판정 tsv 가 들어 있는 폴더")
    ap.add_argument("-o", "--out", type=Path, required=True)
    ap.add_argument("--root", type=Path, default=Path("."))
    ap.add_argument("--map", type=Path, help="doc_id<TAB>deck 파일명 매핑")
    args = ap.parse_args()

    mapping = {r[0]: r[1] for r in read_tsv(args.map)} if args.map else {}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    totals: dict[str, dict[str, int]] = {}

    for tsv in sorted(args.judgments.glob("*.tsv")):
        if tsv.name.endswith(".extra.tsv"):
            continue
        doc_id = tsv.stem
        deck = mapping.get(doc_id)
        if not deck:
            cand = sorted((args.root / "claims").glob(f"{doc_id}__*.jsonl"))
            if not cand:
                print(f"!! {doc_id}: claims 파일을 못 찾았다"); continue
            deck = cand[0].stem

        claims = {c["claim_id"]: c for c in split_claims(args.root / "claims" / f"{deck}.jsonl")}
        sents = load_sents(args.root, doc_id)
        by_order = {i: t for t, i in sents.values()}
        extra = {r[0]: (r[1], r[2] if len(r) > 2 else "") for r in read_tsv(tsv.with_suffix(".extra.tsv"))}

        ws = wb.create_sheet(doc_id)
        for i, name in enumerate(COLUMNS, start=1):
            c = ws.cell(1, i, name)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=HEADER_FILL)
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(i)].width = WIDTHS[i - 1]
        ws.freeze_panes = "C2"

        count = dict.fromkeys(LABEL_FILL, 0)
        for row in read_tsv(tsv):
            cid, label, ref = row[0], row[1], (row[2] if len(row) > 2 else "-")
            claim = claims.get(cid)
            if claim is None:
                print(f"!! {doc_id} {cid}: claim 목록에 없다"); continue
            count[label] = count.get(label, 0) + 1

            core = ctx = where = ""
            if ref in sents:
                core, idx = sents[ref]
                span = [by_order[j] for j in range(idx - CONTEXT, idx + CONTEXT + 1) if j in by_order]
                ctx = " ".join(span)
                where = ref
            elif ref in extra:
                core, note = extra[ref]
                ctx, where = note, "원문 PDF"

            r = ws.max_row + 1
            for i, v in enumerate([cid, claim["slide"], claim["claim"], label, core, ctx, where], start=1):
                cell = ws.cell(r, i, v)
                cell.alignment = Alignment(wrap_text=i in (3, 5, 6), vertical="top")
                cell.border = BORDER
            ws.cell(r, 4).fill = PatternFill("solid", fgColor=LABEL_FILL.get(label, "FFFFFF"))
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"
        totals[doc_id] = count
        print(f"  {doc_id:10s} {ws.max_row-1:4d} claim  " +
              "  ".join(f"{k} {v}" for k, v in count.items() if v))

    s = wb.create_sheet("_요약", 0)
    head = ["덱", "claim", *LABEL_FILL]
    for i, name in enumerate(head, start=1):
        c = s.cell(1, i, name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
        s.column_dimensions[get_column_letter(i)].width = 16
    for doc_id, count in totals.items():
        s.append([doc_id, sum(count.values()), *(count.get(k, 0) for k in LABEL_FILL)])
    s.append(["합계", sum(sum(c.values()) for c in totals.values()),
              *(sum(c.get(k, 0) for c in totals.values()) for k in LABEL_FILL)])

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"\n{args.out}")


if __name__ == "__main__":
    main()
