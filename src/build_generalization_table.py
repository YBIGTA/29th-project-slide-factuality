#!/usr/bin/env python3
"""generalization(재무) 최종 비교표 — M1 · M2 · M2' · M3 · LLM.

    python src/build_generalization_table.py

LLM 판정은 `annotation/finance_claim_annotation_all.xlsx` 에서 읽는다.
원문 PDF 와 덱을 통째로 넣고 받은 결과라 **검색기를 거치지 않는다** — M1~M3 와
입력 조건이 다르다. 표에도 그렇게 적는다.

표를 두 벌 만든다.

  전체 538행   M1~M3 만. LLM 은 44% 만 판정해서 같은 자로 못 잰다.
  공통 239행   LLM 이 판정한 행만 골라 다섯을 다시 집계. 이게 정면 비교다.

두 번째 표를 위해 각 모델의 행별 예측이 필요하다. `eval_generalization.py` 가
결과 json 의 `predictions` 에 남겨 둔다 — 없으면 그 스크립트를 먼저 돌려야 한다.
"""
from __future__ import annotations

import json
import re
import unicodedata
from collections import Counter
from pathlib import Path

import openpyxl
from sklearn.metrics import classification_report, confusion_matrix, f1_score

ROOT = Path(".")
XLSX = ROOT / "annotation" / "finance_claim_annotation_all.xlsx"
DATA = ROOT / "dataset" / "generalization.jsonl"
OUT_MD = ROOT / "results" / "model_compare_generalization.md"
OUT_JSON = ROOT / "results" / "model_compare_generalization.json"

LABELS = ["근거 있음", "무근거", "모순", "Benign"]
OUT_LABELS = ["근거있음", "무근거", "모순", "benign"]
TO_OUT = dict(zip(LABELS, OUT_LABELS))

MODELS = [
    ("M1", "M1 (frozen e5 + LinearSVC)"),
    ("M2", "M2 (KLUE-RoBERTa, 맥락 없음)"),
    ("M2p", "M2' (동일 백본 대조군)"),
    ("M3", "M3 (KLUE-RoBERTa, 맥락 포함)"),
]


def norm(s) -> str:
    s = unicodedata.normalize("NFKC", str(s or ""))
    s = s.translate(str.maketrans({"‘": "'", "’": "'", "“": '"', "”": '"',
                                   "․": "·", "∙": "·", "–": "-", "—": "-"}))
    return re.sub(r"\s+", " ", s).strip()


def nospace(s) -> str:
    return re.sub(r"\s+", "", norm(s))


def canon_label(s) -> str | None:
    return {"근거있음": "근거 있음", "무근거": "무근거", "모순": "모순",
            "benign": "Benign"}.get(norm(s).lower().replace(" ", ""))


def score(gold: list[str], pred: list[str]) -> dict:
    g = [TO_OUT[x] for x in gold]
    p = [TO_OUT[x] for x in pred]
    rep = classification_report(g, p, labels=OUT_LABELS, output_dict=True, zero_division=0)
    return {
        "n": len(g),
        "macro_f1": f1_score(g, p, labels=OUT_LABELS, average="macro", zero_division=0),
        "accuracy": rep["accuracy"],
        "per_class_f1": {l: round(rep[l]["f1-score"], 4) for l in OUT_LABELS},
        "support": {l: int(rep[l]["support"]) for l in OUT_LABELS},
        "confusion": confusion_matrix(g, p, labels=OUT_LABELS).tolist(),
        "confusion_labels": OUT_LABELS,
    }


def load_llm() -> tuple[dict[str, str], list[dict]]:
    """claim_id -> 예측 라벨. 두 번째 값은 라벨을 못 읽은 행."""
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    out: dict[str, str] = {}
    bad: list[dict] = []
    for sheet in [s for s in wb.sheetnames if s != "요약"]:
        for r in list(wb[sheet].iter_rows(values_only=True))[1:]:
            if not r or not r[0]:
                continue
            cid, claim, lab = norm(r[0]), norm(r[1]), canon_label(r[2])
            if lab is None:
                bad.append({"sheet": sheet, "claim_id": cid, "raw_label": str(r[2])})
                continue
            out[cid] = lab
    return out, bad


def main() -> int:
    gen = [json.loads(l) for l in DATA.open(encoding="utf-8") if l.strip()]
    gold_rows = [r for r in gen if isinstance(r.get("label"), str) and r["label"] in LABELS]
    gold = {r["claim_id"]: r["label"] for r in gold_rows}

    llm, bad = load_llm()

    # ── 매칭 ────────────────────────────────────────────────
    by_text = {}
    by_ns = {}
    for r in gold_rows:
        by_text.setdefault(norm(r["claim_text"]), r["claim_id"])
        by_ns.setdefault(nospace(r["claim_text"]), r["claim_id"])

    wb = openpyxl.load_workbook(XLSX, read_only=True)
    claim_of = {}
    for sheet in [s for s in wb.sheetnames if s != "요약"]:
        for r in list(wb[sheet].iter_rows(values_only=True))[1:]:
            if r and r[0]:
                claim_of[norm(r[0])] = norm(r[1])

    llm_matched: dict[str, str] = {}
    unmatched: list[dict] = []
    how = Counter()
    for cid, lab in llm.items():
        if cid in gold:
            llm_matched[cid] = lab
            how["claim_id"] += 1
            continue
        t = claim_of.get(cid, "")
        alt = by_text.get(norm(t)) or by_ns.get(nospace(t))
        if alt and alt not in llm_matched:
            llm_matched[alt] = lab
            how["claim_text" if by_text.get(norm(t)) else "nospace"] += 1
        else:
            unmatched.append({"claim_id": cid, "claim_text": t})

    shared = [cid for cid in gold if cid in llm_matched]
    print(f"LLM {len(llm)}행 -> 매칭 {len(llm_matched)} ({len(llm_matched)/len(llm):.1%}), "
          f"실패 {len(unmatched)}")
    print(f"공통 평가 대상 {len(shared)}행 / 전체 {len(gold)}행 "
          f"({len(shared)/len(gold):.1%})")

    # ── 모델 예측 ──────────────────────────────────────────
    preds: dict[str, dict[str, str]] = {}
    full: dict[str, dict] = {}
    for m, _ in MODELS:
        p = ROOT / "results" / f"{m}_test_everything_543.json"
        d = json.loads(p.read_text(encoding="utf-8"))
        if "predictions" not in d:
            raise SystemExit(f"{p} 에 predictions 가 없다. "
                             "src/eval_generalization.py 를 먼저 다시 돌려라.")
        preds[m] = d["predictions"]
        full[m] = d["runs"][0]["generalization"]

    full["LLM"] = None  # 전체 538행은 평가 불가
    preds["LLM"] = llm_matched

    # ── 집계 ────────────────────────────────────────────────
    sub_gold = [gold[c] for c in shared]
    sub = {m: score(sub_gold, [preds[m][c] for c in shared])
           for m in list(preds)}

    payload = {
        "dataset": {"path": str(DATA), "valid_rows": len(gold),
                    "shared_rows": len(shared),
                    "coverage": len(shared) / len(gold)},
        "llm_source": str(XLSX),
        "matching": {"llm_rows": len(llm), "matched": len(llm_matched),
                     "unmatched": unmatched, "by_method": dict(how),
                     "unreadable_label_rows": bad},
        "full_538": {m: full[m] for m, _ in MODELS},
        "shared_239": sub,
    }
    OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
                        encoding="utf-8")

    # ── 표 ──────────────────────────────────────────────────
    L: list[str] = []
    add = L.append
    n_all, n_sub = len(gold), len(shared)

    add("# generalization(재무) 최종 비교 — M1 · M2 · M2' · M3 · LLM")
    add("")
    add(f"평가 대상은 `dataset/generalization.jsonl` 의 재무 5개 덱이다 "
        f"(`finance_01`~`finance_05`, 라벨 유효 {n_all}행).")
    add("")
    add("**표가 두 벌인 이유.** LLM 은 전체가 아니라 "
        f"{n_sub}행({n_sub / n_all:.0%})만 판정했다. 그래서 모델끼리 비교하는 표와, "
        "LLM 을 포함해 같은 행으로 비교하는 표를 나눠 놓는다. "
        "**두 표의 숫자를 섞어 읽으면 안 된다.**")
    add("")

    # 표 1
    add(f"## 1. 전체 {n_all}행 — 모델 4종")
    add("")
    add("| 모델 | macro-F1 | 정확도 | 근거있음 | 무근거 | 모순 | benign |")
    add("| --- | ---: | ---: | ---: | ---: | ---: | ---: |")
    for m, name in MODELS:
        f = full[m]
        cls = " | ".join(f"{f['per_class_f1'][c]:.3f}" for c in OUT_LABELS)
        add(f"| {name} | **{f['macro_f1']:.4f}** | {f['accuracy']:.4f} | {cls} |")
    add(f"| LLM (원문 직접 판정) | — | — | — | — | — | — |")
    add("")
    add(f"LLM 행이 비어 있는 것은 {n_all}행 중 {n_all - n_sub}행을 판정하지 "
        "않았기 때문이다. 아래 표 2 를 볼 것.")
    add("")
    add("정답 분포: " + " · ".join(
        f"{c} {full['M1']['support'][c]}" for c in OUT_LABELS))
    add("")

    # 표 2
    add(f"## 2. 공통 {n_sub}행 — LLM 포함 정면 비교")
    add("")
    add("LLM 이 판정한 행만 골라 다섯을 **같은 행으로** 다시 집계했다. "
        "모델 쪽은 재추론 없이 저장된 행별 예측에서 뽑았다.")
    add("")
    add("| 모델 | macro-F1 | 정확도 | 근거있음 | 무근거 | 모순 | benign |")
    add("| --- | ---: | ---: | ---: | ---: | ---: | ---: |")
    order = sorted(sub, key=lambda m: -sub[m]["macro_f1"])
    names = dict(MODELS) | {"LLM": "**LLM (원문 직접 판정)**"}
    for m in order:
        s = sub[m]
        cls = " | ".join(f"{s['per_class_f1'][c]:.3f}" for c in OUT_LABELS)
        add(f"| {names[m]} | **{s['macro_f1']:.4f}** | {s['accuracy']:.4f} | {cls} |")
    add("")
    add("정답 분포: " + " · ".join(
        f"{c} {sub['M1']['support'][c]}" for c in OUT_LABELS))
    add("")
    add("```")
    add(f"LLM · 공통 {n_sub}행  (행=정답, 열=예측)")
    add(f"{'':10}" + "".join(f"{c:>9}" for c in OUT_LABELS))
    for c, row in zip(OUT_LABELS, sub["LLM"]["confusion"]):
        add(f"{c:10}" + "".join(f"{v:>9}" for v in row))
    add("```")
    add("")
    pred_cnt = Counter(llm_matched[c] for c in shared)
    add("**LLM 은 'benign' 을 "
        f"{pred_cnt.get('Benign', 0)}번밖에 예측하지 않았는데 정답은 "
        f"{sub['LLM']['support']['benign']}개다.** 슬라이드 제목·목차처럼 "
        "검증 대상이 아닌 텍스트까지 '근거있음' 으로 판정했다. "
        "`guideline_v2.md` 는 고유명사·소제목을 Benign 으로 빼지 말라고 하고, "
        "`results/iaa_v1.md` 는 이 기준이 어노테이터 사이에서도 가장 크게 "
        "갈렸다고 기록한다(불일치 유형 1). 이 표의 LLM 점수는 모델 성능만큼이나 "
        "**Benign 기준 해석 차이**를 재고 있다.")
    add("")

    # 각주
    add("## 읽을 때 주의할 것")
    add("")
    add("1. **LLM 은 입력 조건이 다르다.** 원문 PDF 와 덱을 통째로 넣고 받은 판정이라 "
        "**검색기를 거치지 않는다.** M1~M3 는 검색기가 뽑은 top-5 구간만 보고 "
        "판정한다. 즉 이 비교는 '모델 대 LLM' 인 동시에 "
        "'검색된 근거만 보기 대 원문 전체 보기' 다. LLM 이 유리한 조건이다.")
    add("")
    add(f"2. **커버리지 {n_sub}/{n_all} ({n_sub / n_all:.0%}).** "
        "빠진 행이 무작위가 아니다. 정답 '모순' "
        f"{full['M1']['support']['모순']}개 중 "
        f"{full['M1']['support']['모순'] - sub['M1']['support']['모순']}개를 "
        "LLM 이 보지 않았다. 가장 어려운 클래스가 대거 빠져 있어 "
        "LLM 쪽에 유리한 방향으로 치우친다.")
    add("")
    add(f"3. **매칭 실패 {len(unmatched)}행** "
        f"({len(unmatched) / len(llm):.1%}). claim_id 로 붙였고 실패분은 "
        "claim 본문·공백무시 비교까지 시도했지만 데이터셋에 해당 claim 이 없다.")
    for u in unmatched:
        add(f"   - `{u['claim_id']}` — {u['claim_text'][:60]}")
    add("")
    add("4. **모델 쪽은 단일 seed 다.** 가중치가 모델당 한 벌뿐이라 "
        "3-seed 평균·표준편차를 낼 수 없다. PR #10 의 259행 결과"
        "(seed 13/42/777 평균)와 직접 비교하면 안 된다.")
    add("")
    add("5. **M2 의 입력 조립은 가정이다.** 학습 코드가 레포에 없어 M3 형식으로 "
        "가정해 돌렸다. 실제 학습 형식과 다를 수 있어 **M2 수치는 보증되지 않는다.**")
    add("")
    add("6. **`finance_03`·`finance_05` 의 원문 정제는 사람 검수 전이다** "
        "(종결어미 49.7% / 54.9%). `finance_04` 는 읽기레이어 추출로 75.7%.")
    add("")
    add("7. **후보 구간이 비어 있는 행 17개**(`finance_03` 9 · `finance_05` 8)는 "
        "모델이 근거 없이 판정한다. LLM 은 원문을 직접 보므로 이 제약을 받지 않는다.")
    add("")
    add("---")
    add("")
    add("생성: `python src/build_generalization_table.py` · "
        "수치 원본은 `results/model_compare_generalization.json`")

    OUT_MD.write_text("\n".join(L) + "\n", encoding="utf-8")
    print(f"-> {OUT_MD}")
    print(f"-> {OUT_JSON}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
