from __future__ import annotations

"""Phase C M1: 고정된 multilingual-E5 임베딩 + 경량 분류기.

데이터 분할 기준은 ``annotation/공통_덱_annotation_정리본.xlsx``에서 읽는다.

* train: finance가 아니면서 ``Test Set`` 값이 1이 아닌 덱
* test: finance가 아니면서 ``Test Set`` 값이 1로 시작하는 덱
* generalization: 모든 finance 덱

하이퍼파라미터 선택과 모델 학습에는 train split만 사용한다. 보류해 둔 test
split과 generalization split은 분류기가 확정된 뒤 임베딩하고 평가한다.
"""

import argparse
import hashlib
import json
import platform
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import joblib
import numpy as np
from openpyxl import load_workbook
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, classification_report, confusion_matrix, f1_score
from sklearn.model_selection import GroupKFold
from sklearn.svm import SVC, LinearSVC


MODEL_ID = "M1"
DEFAULT_ENCODER = "intfloat/multilingual-e5-large"
LABELS = ["근거 있음", "무근거", "모순", "Benign"]
LABEL_TO_ID = {label: i for i, label in enumerate(LABELS)}
FEATURE_SPEC = "[claim, best_candidate, abs_diff, product, max/mean/std/min_similarity, candidate_count]"


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8") as stream:
        return [json.loads(line) for line in stream if line.strip()]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_deck_id(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.split("__", 1)[0]
    text = re.sub(r"\s+", "", text)
    if text.startswith("fin_"):
        text = "finance_" + text.removeprefix("fin_")
    return text


def is_test_marker(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return float(value) == 1.0
    return bool(re.match(r"^\s*1(?:\D|$)", str(value)))


def resolve_split_decks(annotation_path: Path) -> dict[str, list[str]]:
    """인덱스 시트에서 중복을 제거한 train/test/generalization 덱 ID를 구한다."""
    workbook = load_workbook(annotation_path, read_only=True, data_only=True)
    if "_시트이름" not in workbook.sheetnames:
        raise ValueError(f"{annotation_path}: '_시트이름' sheet is missing")
    sheet = workbook["_시트이름"]
    headers = {str(cell.value).strip(): i for i, cell in enumerate(next(sheet.iter_rows()))}
    required = {"deck_id", "레포 파일명"}
    missing = required - headers.keys()
    if missing:
        raise ValueError(f"{annotation_path}: missing columns {sorted(missing)}")
    if "split" not in headers and "Test Set" not in headers:
        raise ValueError(f"{annotation_path}: missing 'split' or 'Test Set' column")

    buckets: dict[str, set[str]] = {"train": set(), "test": set(), "generalization": set()}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        deck = normalize_deck_id(row[headers["레포 파일명"]] or row[headers["deck_id"]])
        if not deck:
            continue

        if "split" in headers:
            split = str(row[headers["split"]] or "").strip().lower()
            if split not in buckets:
                raise ValueError(f"{annotation_path}: invalid split {split!r} for {deck}")
        else:
            marker = row[headers["Test Set"]]
            marker_text = str(marker or "").lower().replace("_", "")
            if "tech4" in marker_text:
                deck = "tech_04"
            if deck.startswith("finance_"):
                split = "generalization"
            elif is_test_marker(marker):
                split = "test"
            else:
                split = "train"
        buckets[split].add(deck)

    workbook.close()
    overlap = (buckets["train"] & buckets["test"]) | (buckets["train"] & buckets["generalization"]) | (
        buckets["test"] & buckets["generalization"]
    )
    if overlap:
        raise ValueError(f"deck IDs assigned to multiple splits: {sorted(overlap)}")
    return {name: sorted(values) for name, values in buckets.items()}


def valid_rows(rows: Iterable[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    kept: list[dict[str, Any]] = []
    invalid_by_deck: Counter[str] = Counter()
    invalid_values: Counter[str] = Counter()
    for row in rows:
        if row.get("label") in LABEL_TO_ID:
            kept.append(row)
        else:
            invalid_by_deck[str(row.get("deck_id", ""))] += 1
            invalid_values[repr(row.get("label"))] += 1
    return kept, {
        "count": sum(invalid_by_deck.values()),
        "by_deck": dict(sorted(invalid_by_deck.items())),
        "raw_values": dict(sorted(invalid_values.items())),
    }


def load_selected_rows(dataset_dir: Path, annotation_path: Path) -> tuple[
    list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]
]:
    split_decks = resolve_split_decks(annotation_path)
    source_paths = [dataset_dir / name for name in ("train.jsonl", "test.jsonl", "generalization.jsonl")]
    all_rows = [row for path in source_paths for row in read_jsonl(path)]

    by_deck: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in all_rows:
        by_deck[normalize_deck_id(row.get("deck_id") or row.get("doc_id"))].append(row)

    available = set(by_deck)
    missing_train = set(split_decks["train"]) - available
    missing_test = set(split_decks["test"]) - available
    missing_generalization = set(split_decks["generalization"]) - available
    if missing_train or missing_test:
        raise ValueError(
            "dataset rows missing for annotation split: "
            f"train={sorted(missing_train)}, test={sorted(missing_test)}"
        )

    train_raw = [row for deck in split_decks["train"] for row in by_deck[deck]]
    test_raw = [row for deck in split_decks["test"] for row in by_deck[deck]]
    generalization_raw = [row for deck in split_decks["generalization"] for row in by_deck[deck]]
    train_rows, train_invalid = valid_rows(train_raw)
    test_rows, test_invalid = valid_rows(test_raw)
    generalization_rows, generalization_invalid = valid_rows(generalization_raw)
    evaluated_generalization_decks = sorted(
        {normalize_deck_id(row.get("deck_id") or row.get("doc_id")) for row in generalization_rows}
    )

    leaked = [r.get("claim_id") for r in train_rows + test_rows if normalize_deck_id(r.get("deck_id")).startswith("finance_")]
    if leaked:
        raise RuntimeError(f"finance leakage detected in M1 train/test: {leaked[:5]}")
    if not train_rows or not test_rows or not generalization_rows:
        raise ValueError("train, test, and generalization must all contain at least one valid row")

    metadata = {
        "split_decks": split_decks,
        "source_files": {
            str(path): {"sha256": sha256_file(path), "rows": len(read_jsonl(path))}
            for path in source_paths
        },
        "annotation": {"path": str(annotation_path), "sha256": sha256_file(annotation_path)},
        "raw_counts": {
            "train": len(train_raw),
            "test": len(test_raw),
            "generalization": len(generalization_raw),
        },
        "valid_counts": {
            "train": len(train_rows),
            "test": len(test_rows),
            "generalization": len(generalization_rows),
        },
        "invalid_labels": {
            "train": train_invalid,
            "test": test_invalid,
            "generalization": generalization_invalid,
        },
        "missing_generalization_decks": sorted(missing_generalization),
        "evaluated_generalization_decks": evaluated_generalization_decks,
    }
    return train_rows, test_rows, generalization_rows, metadata


def candidate_texts(row: dict[str, Any]) -> list[str]:
    return [
        str(candidate.get("text", "")).strip()
        for candidate in (row.get("candidates") or [])[:5]
        if str(candidate.get("text", "")).strip()
    ]


def encode_lookup(model: Any, texts: list[str], prefix: str, batch_size: int) -> dict[str, np.ndarray]:
    unique = list(dict.fromkeys(texts))
    if not unique:
        return {}
    vectors = model.encode(
        [f"{prefix}{text}" for text in unique],
        batch_size=batch_size,
        normalize_embeddings=True,
        convert_to_numpy=True,
        show_progress_bar=True,
    )
    return {text: np.asarray(vector, dtype=np.float32) for text, vector in zip(unique, vectors)}


def build_features(
    rows: list[dict[str, Any]], claim_vectors: dict[str, np.ndarray], passage_vectors: dict[str, np.ndarray], dimension: int
) -> np.ndarray:
    features: list[np.ndarray] = []
    zero = np.zeros(dimension, dtype=np.float32)
    for row in rows:
        claim = str(row.get("claim_text", "")).strip()
        q = claim_vectors[claim]
        texts = candidate_texts(row)
        if texts:
            candidates = np.stack([passage_vectors[text] for text in texts])
            similarities = candidates @ q
            best = candidates[int(np.argmax(similarities))]
            statistics = np.asarray(
                [similarities.max(), similarities.mean(), similarities.std(), similarities.min(), len(texts) / 5.0],
                dtype=np.float32,
            )
        else:
            best = zero
            statistics = np.zeros(5, dtype=np.float32)
        features.append(np.concatenate([q, best, np.abs(q - best), q * best, statistics]))
    return np.stack(features).astype(np.float32, copy=False)


def class_counts(rows: list[dict[str, Any]]) -> dict[str, int]:
    counts = Counter(str(row["label"]) for row in rows)
    return {label: counts.get(label, 0) for label in LABELS}


def classifier_candidates() -> list[dict[str, Any]]:
    """성능 가능성이 높은 세 경량 분류기의 비교 조합을 만든다."""
    candidates: list[dict[str, Any]] = []
    for class_weight in (None, "balanced"):
        for c_value in (0.1, 1.0, 10.0, 100.0):
            candidates.append({"type": "LogisticRegression", "C": c_value, "class_weight": class_weight})
        for c_value in (0.01, 0.1, 1.0, 10.0):
            candidates.append({"type": "LinearSVC", "C": c_value, "class_weight": class_weight})
        for c_value in (1.0, 10.0, 100.0):
            candidates.append({"type": "RbfSVC", "C": c_value, "class_weight": class_weight})
    return candidates


def make_classifier(specification: dict[str, Any], seed: int) -> Any:
    """후보 명세 하나를 실제 scikit-learn 분류기로 변환한다."""
    common = {
        "C": specification["C"],
        "class_weight": specification["class_weight"],
        "random_state": seed,
    }
    if specification["type"] == "LogisticRegression":
        return LogisticRegression(**common, max_iter=3000, solver="lbfgs")
    if specification["type"] == "LinearSVC":
        return LinearSVC(**common, dual="auto", max_iter=5000)
    if specification["type"] == "RbfSVC":
        return SVC(**common, kernel="rbf", gamma="scale", cache_size=1024)
    raise ValueError(f"지원하지 않는 분류기: {specification['type']}")


def tune_classifier(
    features: np.ndarray, targets: np.ndarray, groups: np.ndarray, seed: int
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    """Test를 보지 않고 train 덱 단위 교차검증으로 최적 분류기를 선택한다."""
    unique_groups = np.unique(groups)
    folds = min(5, len(unique_groups))
    if folds < 2:
        raise ValueError("at least two train decks are required for group cross-validation")
    splitter = GroupKFold(n_splits=folds)
    trials: list[dict[str, Any]] = []
    for specification in classifier_candidates():
        scores: list[float] = []
        for train_idx, validation_idx in splitter.split(features, targets, groups):
            classifier = make_classifier(specification, seed)
            classifier.fit(features[train_idx], targets[train_idx])
            prediction = classifier.predict(features[validation_idx])
            scores.append(float(f1_score(targets[validation_idx], prediction, labels=range(len(LABELS)), average="macro", zero_division=0)))
        trials.append({
            **specification,
            "fold_macro_f1": scores,
            "mean_macro_f1": float(np.mean(scores)),
            "std_macro_f1": float(np.std(scores)),
        })

    # 평균이 같으면 fold 간 편차가 작은 모델을, 그것도 같으면 더 단순한 모델을 택한다.
    simplicity = {"LogisticRegression": 0, "LinearSVC": 1, "RbfSVC": 2}
    best = min(
        trials,
        key=lambda trial: (
            -trial["mean_macro_f1"],
            trial["std_macro_f1"],
            simplicity[trial["type"]],
            trial["C"],
        ),
    )
    return {key: best[key] for key in ("type", "C", "class_weight")}, trials


def per_deck_metrics(rows: list[dict[str, Any]], targets: np.ndarray, predictions: np.ndarray) -> dict[str, Any]:
    indices: dict[str, list[int]] = defaultdict(list)
    for i, row in enumerate(rows):
        indices[normalize_deck_id(row.get("deck_id"))].append(i)
    output: dict[str, Any] = {}
    for deck, idx in sorted(indices.items()):
        y_true = targets[idx]
        y_pred = predictions[idx]
        output[deck] = {
            "support": len(idx),
            "accuracy": float(accuracy_score(y_true, y_pred)),
            "macro_f1_all_labels": float(f1_score(y_true, y_pred, labels=range(len(LABELS)), average="macro", zero_division=0)),
        }
    return output


def write_json(path: Path, value: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(value, ensure_ascii=False, indent=2, allow_nan=False) + "\n", encoding="utf-8")
    temporary.replace(path)


def write_readme(
    path: Path,
    encoder: str,
    train_count: int,
    test_count: int,
    generalization_count: int,
    classifier_specification: dict[str, Any],
) -> None:
    text = f"""# M1 - frozen embedding + lightweight classifier

- Encoder: `{encoder}` (frozen; no encoder parameters are updated)
- Classifier: scikit-learn `{classifier_specification['type']}` (`class_weight={classifier_specification['class_weight']}`, `C={classifier_specification['C']}`)
- Input: claim and up to five retrieved candidate passages; slide context is excluded to match M2's no-context condition
- Train: non-finance decks with `Test Set != 1` ({train_count} valid claims)
- Test: non-finance decks with `Test Set = 1` ({test_count} valid claims)
- Generalization: every finance deck ({generalization_count} valid claims; evaluation only)

`classifier.joblib` contains the fitted classifier, label mapping, and feature specification.
`config.json` records the exact deck split and SHA-256 hashes of every input file.
The frozen encoder is referenced by Hugging Face model ID rather than duplicated in this directory.

Reproduce from the repository root:

```bash
python src/train_m1.py --local-files-only
```
"""
    path.write_text(text, encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--root", type=Path, default=Path("."))
    parser.add_argument("--annotation", type=Path, default=Path("annotation/공통_덱_annotation_정리본.xlsx"))
    parser.add_argument("--dataset-dir", type=Path, default=Path("dataset"))
    parser.add_argument("--model-dir", type=Path, default=Path("models/M1"))
    parser.add_argument("--result", type=Path, default=Path("results/M1_test.json"))
    parser.add_argument("--encoder", default=DEFAULT_ENCODER)
    parser.add_argument("--device", default="cpu")
    parser.add_argument("--batch-size", type=int, default=16)
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--local-files-only", action="store_true")
    args = parser.parse_args()

    root = args.root.resolve()
    annotation = (root / args.annotation).resolve() if not args.annotation.is_absolute() else args.annotation
    dataset_dir = (root / args.dataset_dir).resolve() if not args.dataset_dir.is_absolute() else args.dataset_dir
    model_dir = (root / args.model_dir).resolve() if not args.model_dir.is_absolute() else args.model_dir
    result_path = (root / args.result).resolve() if not args.result.is_absolute() else args.result

    train_rows, test_rows, generalization_rows, metadata = load_selected_rows(dataset_dir, annotation)
    print("split decks:", json.dumps(metadata["split_decks"], ensure_ascii=False))
    print("valid rows:", metadata["valid_counts"], "invalid labels:", metadata["invalid_labels"])

    from sentence_transformers import SentenceTransformer
    import sentence_transformers
    import sklearn
    import torch

    encoder = SentenceTransformer(
        args.encoder,
        device=args.device,
        local_files_only=args.local_files_only,
    )
    dimension = int(encoder.get_embedding_dimension())
    all_rows = train_rows + test_rows + generalization_rows
    claims = [str(row.get("claim_text", "")).strip() for row in all_rows]
    passages = [text for row in all_rows for text in candidate_texts(row)]
    print(f"encoding {len(set(claims))} unique claims and {len(set(passages))} unique candidate passages")
    claim_vectors = encode_lookup(encoder, claims, "query: ", args.batch_size)
    passage_vectors = encode_lookup(encoder, passages, "passage: ", args.batch_size)
    train_x = build_features(train_rows, claim_vectors, passage_vectors, dimension)
    test_x = build_features(test_rows, claim_vectors, passage_vectors, dimension)
    generalization_x = build_features(generalization_rows, claim_vectors, passage_vectors, dimension)
    train_y = np.asarray([LABEL_TO_ID[row["label"]] for row in train_rows], dtype=np.int64)
    test_y = np.asarray([LABEL_TO_ID[row["label"]] for row in test_rows], dtype=np.int64)
    generalization_y = np.asarray(
        [LABEL_TO_ID[row["label"]] for row in generalization_rows], dtype=np.int64
    )
    train_groups = np.asarray([normalize_deck_id(row.get("deck_id")) for row in train_rows])

    best_specification, cv_trials = tune_classifier(train_x, train_y, train_groups, args.seed)
    classifier = make_classifier(best_specification, args.seed)
    classifier.fit(train_x, train_y)
    prediction = classifier.predict(test_x)
    generalization_prediction = classifier.predict(generalization_x)
    description = (
        "frozen multilingual-E5 embeddings + "
        f"train-selected {best_specification['type']}, no slide context"
    )

    model_dir.mkdir(parents=True, exist_ok=True)
    joblib.dump(
        {
            "classifier": classifier,
            "model_id": MODEL_ID,
            "encoder": args.encoder,
            "label_to_id": LABEL_TO_ID,
            "labels": LABELS,
            "feature_spec": FEATURE_SPEC,
            "embedding_dimension": dimension,
            "classifier_specification": best_specification,
        },
        model_dir / "classifier.joblib",
        compress=3,
    )
    created_at = datetime.now(timezone.utc).isoformat()
    config = {
        "model_id": MODEL_ID,
        "description": description,
        "created_at_utc": created_at,
        "encoder": {"model_id": args.encoder, "frozen": True, "dimension": dimension, "prefixes": {"claim": "query: ", "candidate": "passage: "}},
        "classifier": {**best_specification, "random_state": args.seed},
        "feature_spec": FEATURE_SPEC,
        "labels": LABELS,
        "data": metadata,
        "class_counts": {
            "train": class_counts(train_rows),
            "test": class_counts(test_rows),
            "generalization": class_counts(generalization_rows),
        },
        "cross_validation": {
            "type": "GroupKFold",
            "group": "deck_id",
            "selection_metric": "mean_macro_f1",
            "test_labels_used_for_selection": False,
            "note": "이전 M1 반복에서 test 성능이 이미 관찰되었으므로 최종 test 수치는 완전히 미노출된 평가가 아니다.",
            "trials": cv_trials,
        },
        "environment": {"python": sys.version.split()[0], "platform": platform.platform(), "torch": torch.__version__, "sentence_transformers": sentence_transformers.__version__, "scikit_learn": sklearn.__version__},
    }
    write_json(model_dir / "config.json", config)
    write_readme(
        model_dir / "M1_README.md",
        args.encoder,
        len(train_rows),
        len(test_rows),
        len(generalization_rows),
        best_specification,
    )

    report = classification_report(test_y, prediction, labels=range(len(LABELS)), target_names=LABELS, output_dict=True, zero_division=0)
    generalization_report = classification_report(
        generalization_y,
        generalization_prediction,
        labels=range(len(LABELS)),
        target_names=LABELS,
        output_dict=True,
        zero_division=0,
    )
    result = {
        "model_id": MODEL_ID,
        "description": description,
        "selected_classifier": best_specification,
        "evaluation_split": "non-finance annotation decks with Test Set = 1",
        "test_decks": metadata["split_decks"]["test"],
        "generalization_decks": metadata["evaluated_generalization_decks"],
        "train_support": len(train_rows),
        "test_support": len(test_rows),
        "macro_f1": float(f1_score(test_y, prediction, labels=range(len(LABELS)), average="macro", zero_division=0)),
        "accuracy": float(accuracy_score(test_y, prediction)),
        "classification_report": report,
        "confusion_matrix": confusion_matrix(test_y, prediction, labels=range(len(LABELS))).tolist(),
        "label_order": LABELS,
        "per_deck": per_deck_metrics(test_rows, test_y, prediction),
        "generalization": {
            "evaluation_split": "finance generalization decks",
            "support": len(generalization_rows),
            "macro_f1": float(
                f1_score(
                    generalization_y,
                    generalization_prediction,
                    labels=range(len(LABELS)),
                    average="macro",
                    zero_division=0,
                )
            ),
            "accuracy": float(accuracy_score(generalization_y, generalization_prediction)),
            "classification_report": generalization_report,
            "confusion_matrix": confusion_matrix(
                generalization_y, generalization_prediction, labels=range(len(LABELS))
            ).tolist(),
            "per_deck": per_deck_metrics(
                generalization_rows, generalization_y, generalization_prediction
            ),
        },
        "invalid_labels_excluded": metadata["invalid_labels"],
        "model_config": str((model_dir / "config.json").relative_to(root)),
    }
    write_json(result_path, result)
    print(f"saved {model_dir}")
    print(
        f"saved {result_path}: test macro-F1={result['macro_f1']:.4f}, "
        f"generalization macro-F1={result['generalization']['macro_f1']:.4f}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
