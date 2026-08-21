#!/usr/bin/env python3
"""정리본 워크북의 각 시트에 담당자·deck_id 를 컬럼으로 박고,
요약 시트에 원문·덱 자료가 실제로 레포에 있는지를 붙인다.

    python src/stamp_annotation.py

시트 이름(`arts_01__시나`)에만 담당자가 들어 있으면 시트를 CSV 로 뽑거나
행을 다른 표에 붙여넣는 순간 누가 단 라벨인지가 사라진다. 컬럼으로 박아둔다.

같은 파일을 여러 번 돌려도 안전하다 — 이미 박혀 있으면 값만 갱신한다.
"""
from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NEW_COLS = ["deck_id", "담당자"]          # 시트 맨 앞에 넣는다
WIDTHS = {"deck_id": 22, "담당자": 10}
SUMMARY = "_시트이름"

# 요약 시트에 덧붙이는 자료 유무 컬럼
ASSET_COLS = ["원문 PDF", "정제 텍스트", "덱 PPTX", "manifest"]

HEADER_FILL = "355BB7"
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MISSING_FILL = "FDB9C3"     # X 인 칸 — 정리본 라벨 색과 같은 분홍


def header_cell(cell, text):
    cell.value = text
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def split_sheet_name(name: str) -> tuple[str, str]:
    """`arts_01__시나`, `tech_03__지원(중복)` -> ('arts_01', '시나')"""
    doc, _, who = name.partition("__")
    return doc, re.sub(r"\(.*\)$", "", who).strip()


def assets(root: Path) -> dict[str, dict[str, str]]:
    """doc_id 별로 레포에 무엇이 있는지."""
    raw = {p.stem for p in (root / "docs" / "raw").glob("*.pdf")}
    clean = {p.stem for p in (root / "docs" / "clean").glob("*.txt")
             if not p.stem.endswith(".review")}
    # 덱 파일명은 `{doc_id}__{생성기}` 라 doc_id 접두사로 찾는다
    decks = {p.stem.split("__")[0] for p in (root / "decks").glob("*.pptx")}
    man: set[str] = set()
    mpath = root / "docs" / "manifest.csv"
    if mpath.exists():
        # utf-8-sig — 엑셀이 붙인 BOM 때문에 첫 컬럼명이 어긋난다
        with mpath.open(encoding="utf-8-sig", newline="") as f:
            man = {r["doc_id"] for r in csv.DictReader(f) if r.get("doc_id")}
    out = {}
    for d in raw | clean | decks | man:
        out[d] = {"원문 PDF": "O" if d in raw else "X",
                  "정제 텍스트": "O" if d in clean else "X",
                  "덱 PPTX": "O" if d in decks else "X",
                  "manifest": "O" if d in man else "X"}
    return out


def stamp_sheet(ws, deck_id: str, who: str) -> int:
    """맨 앞에 deck_id·담당자 컬럼을 넣고 전 행에 값을 채운다."""
    have = [ws.cell(1, i + 1).value for i in range(len(NEW_COLS))]
    if have != NEW_COLS:
        ws.insert_cols(1, len(NEW_COLS))
    for i, name in enumerate(NEW_COLS, start=1):
        header_cell(ws.cell(1, i), name)
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS[name]

    n = 0
    for r in range(2, ws.max_row + 1):
        # 빈 꼬리 행은 건너뛴다 (오른쪽 데이터 칸이 전부 비었으면 행이 아니다)
        if all(ws.cell(r, c).value in (None, "")
               for c in range(len(NEW_COLS) + 1, ws.max_column + 1)):
            continue
        for i, val in enumerate((deck_id, who), start=1):
            cell = ws.cell(r, i)
            cell.value = val
            cell.alignment = Alignment(vertical="top")
            cell.border = BORDER
        n += 1

    ws.freeze_panes = ws.cell(2, len(NEW_COLS) + 2).coordinate
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    return n


def stamp_summary(ws, table: dict[str, dict[str, str]]) -> None:
    """요약 시트에 자료 유무를 덧붙인다. deck_id 컬럼을 열쇠로 쓴다."""
    head = [c.value for c in ws[1]]
    key = head.index("deck_id") + 1
    start = ws.max_column + 1
    if head[-len(ASSET_COLS):] == ASSET_COLS:
        start = ws.max_column - len(ASSET_COLS) + 1
    for i, name in enumerate(ASSET_COLS):
        header_cell(ws.cell(1, start + i), name)
        ws.column_dimensions[get_column_letter(start + i)].width = 12

    for r in range(2, ws.max_row + 1):
        doc = ws.cell(r, key).value
        if not doc:
            continue
        row = table.get(str(doc), dict.fromkeys(ASSET_COLS, "X"))
        for i, name in enumerate(ASSET_COLS):
            cell = ws.cell(r, start + i)
            cell.value = row[name]
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
            if row[name] == "X":
                cell.fill = PatternFill("solid", fgColor=MISSING_FILL)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("workbook", nargs="?",
                    default="annotation/공통_덱_annotation_정리본.xlsx", type=Path)
    ap.add_argument("--root", type=Path, default=Path("."))
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.workbook)
    table = assets(args.root)

    total = 0
    for name in wb.sheetnames:
        if name == SUMMARY:
            continue
        deck_id, who = split_sheet_name(name)
        n = stamp_sheet(wb[name], deck_id, who)
        total += n
        print(f"  {name:24s} {n:4d}행  <- {deck_id} / {who}")

    if SUMMARY in wb.sheetnames:
        stamp_summary(wb[SUMMARY], table)
        print(f"  {SUMMARY:24s} 자료 유무 {len(ASSET_COLS)}칸 추가")

    wb.save(args.workbook)
    print(f"\n{args.workbook} · 시트 {len(wb.sheetnames)-1}개 · {total}행에 담당자 기입")


if __name__ == "__main__":
    main()
